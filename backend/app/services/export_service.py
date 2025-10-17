"""Excel出力サービス"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Any
import os


class ExportService:
    """Excel出力サービス"""

    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    def write_excel_output(
        self,
        original_excel_path: str,
        estimates: List[Dict[str, Any]],
        totals: Dict[str, float],
        qa_pairs: List[Dict[str, str]],
        output_dir: str,
    ) -> str:
        """Excel形式で見積り結果を出力する"""

        # 出力ファイル名を生成
        output_filename = f"{self.timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        try:
            # 元のExcelファイルを読み込み
            df = pd.read_excel(original_excel_path, engine="openpyxl")

            # 見積りデータを追加
            df = self._add_estimate_columns(df, estimates)

            # Excelファイルを作成
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="見積り", index=False)

                # ワークシートを取得
                worksheet = writer.sheets["見積り"]

                # 書式設定
                self._format_worksheet(worksheet, len(df))

                # 合計情報を追加
                self._add_totals_to_worksheet(worksheet, totals, len(df))

                # セッション情報を追加
                self._add_session_info_to_worksheet(
                    worksheet, qa_pairs, len(df)
                )

            print(f"\n見積り結果をExcelファイルに出力しました: {output_path}")
            return output_path

        except Exception as e:
            print(f"Excelファイル出力でエラーが発生しました: {e}")
            raise

    def _add_estimate_columns(
        self, df: pd.DataFrame, estimates: List[Dict[str, Any]]
    ) -> pd.DataFrame:
        """見積り列（C列：工数、D列：金額、E列：工数内訳、F列：根拠・備考）を追加する"""

        # 見積りデータを辞書として整理
        estimate_dict = {est["name"]: est for est in estimates}

        # 新しい列を追加
        person_days_list = []
        amounts_list = []
        breakdown_list = []
        notes_list = []

        for index, row in df.iterrows():
            deliverable_name = str(row.iloc[0])

            if deliverable_name in estimate_dict:
                estimate = estimate_dict[deliverable_name]
                person_days_list.append(estimate["person_days"])
                amounts_list.append(estimate["amount"])
                breakdown_list.append(estimate.get("reasoning_breakdown", estimate.get("reasoning", "")))
                notes_list.append(estimate.get("reasoning_notes", ""))
            else:
                person_days_list.append(0)
                amounts_list.append(0)
                breakdown_list.append("")
                notes_list.append("")

        # 列を追加
        df["予想工数（人日）"] = person_days_list
        df["金額"] = amounts_list
        df["工数内訳"] = breakdown_list
        df["根拠・備考"] = notes_list

        return df

    def _format_worksheet(self, worksheet, data_rows: int):
        """ワークシートの書式設定を行う"""

        # ヘッダー行の書式設定
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        for col in range(1, 7):  # A列からF列まで
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment

        # データ行の書式設定
        for row in range(2, data_rows + 2):
            # 工数列（C列）の書式設定
            cell = worksheet.cell(row=row, column=3)
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")

            # 金額列（D列）の書式設定
            cell = worksheet.cell(row=row, column=4)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

            # 工数内訳列（E列）の書式設定
            cell = worksheet.cell(row=row, column=5)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # 根拠・備考列（F列）の書式設定
            cell = worksheet.cell(row=row, column=6)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 列幅の調整
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 50
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 15
        worksheet.column_dimensions["E"].width = 40  # 工数内訳
        worksheet.column_dimensions["F"].width = 40  # 根拠・備考

    def _add_totals_to_worksheet(
        self, worksheet, totals: Dict[str, float], data_rows: int
    ):
        """合計情報をワークシートに追加する"""

        # 合計情報を追加する開始行
        start_row = data_rows + 3

        # 小計
        worksheet.cell(row=start_row, column=3, value="小計")
        worksheet.cell(row=start_row, column=4, value=totals["subtotal"])

        # 税額
        worksheet.cell(row=start_row + 1, column=3, value="税額 (10%)")
        worksheet.cell(row=start_row + 1, column=4, value=totals["tax"])

        # 総額
        worksheet.cell(row=start_row + 2, column=3, value="総額")
        worksheet.cell(row=start_row + 2, column=4, value=totals["total"])

        # 合計行の書式設定
        for row in range(start_row, start_row + 3):
            # ラベル列（C列）
            cell = worksheet.cell(row=row, column=3)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")

            # 金額列（D列）
            cell = worksheet.cell(row=row, column=4)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            if row == start_row + 2:  # 総額行
                cell.font = Font(bold=True)
                cell.border = Border(top=Side(style="thin"))

    def _add_session_info_to_worksheet(
        self, worksheet, qa_pairs: List[Dict[str, str]], data_rows: int
    ):
        """セッション情報をワークシートに追加する"""

        # セッション情報を追加する開始行
        start_row = data_rows + 7

        # 質問と回答
        worksheet.cell(row=start_row, column=1, value="【見積り精度向上のための質問と回答】")
        worksheet.cell(row=start_row, column=1).font = Font(bold=True)

        current_row = start_row + 1

        for i, qa in enumerate(qa_pairs, 1):
            worksheet.cell(row=current_row, column=1, value=f"質問{i}")
            worksheet.cell(row=current_row, column=2, value=qa["question"])
            current_row += 1

            worksheet.cell(row=current_row, column=1, value=f"回答{i}")
            worksheet.cell(row=current_row, column=2, value=qa["answer"])
            current_row += 1

            # 空行
            current_row += 1

        # 出力日時
        worksheet.cell(row=current_row + 1, column=1, value="【出力日時】")
        worksheet.cell(row=current_row + 1, column=1).font = Font(bold=True)
        worksheet.cell(
            row=current_row + 1,
            column=2,
            value=datetime.now().strftime("%Y年%m月%d日 %H時%M分"),
        )
