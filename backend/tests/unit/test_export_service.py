"""Unit tests for ExportService"""
import pytest
import pandas as pd
from openpyxl import load_workbook
from app.services.export_service import ExportService


class TestExportService:
    """Test class for ExportService"""

    def test_export_to_excel_success(self, tmp_path, sample_excel_file, mock_language_ja):
        """Test successful Excel export"""
        # Prepare test data
        estimates = [
            {
                "name": "要件定義書",
                "description": "システム全体の要件を定義する文書",
                "person_days": 5.0,
                "amount": 500000.0,
                "reasoning_breakdown": "要件: 2人日\n設計: 3人日",
                "reasoning": "過去実績参照"
            }
        ]

        totals = {"subtotal": 500000.0, "tax": 50000.0, "total": 550000.0}

        qa_pairs = [
            {"question": "ユーザー数は？", "answer": "100名"}
        ]

        # Export to Excel
        service = ExportService()
        output_path = str(tmp_path / "test_export.xlsx")
        result_path = service.write_excel_output(
            sample_excel_file,
            estimates,
            totals,
            qa_pairs,
            str(tmp_path)
        )

        # Verify file was created
        assert result_path.endswith(".xlsx")

        # Verify Excel content
        wb = load_workbook(result_path)
        assert len(wb.sheetnames) > 0

        ws = wb.active
        assert ws is not None

    def test_export_with_multiple_estimates(self, tmp_path, sample_excel_file, mock_language_ja):
        """Test Excel export with multiple estimates"""
        # Prepare test data with multiple estimates
        estimates = [
            {"name": "要件定義書", "description": "要件定義", "person_days": 5.0, "amount": 500000.0, "reasoning": ""},
            {"name": "基本設計書", "description": "基本設計", "person_days": 10.0, "amount": 1000000.0, "reasoning": ""},
            {"name": "詳細設計書", "description": "詳細設計", "person_days": 15.0, "amount": 1500000.0, "reasoning": ""}
        ]

        totals = {"subtotal": 3000000.0, "tax": 300000.0, "total": 3300000.0}
        qa_pairs = []

        service = ExportService()
        result_path = service.write_excel_output(
            sample_excel_file,
            estimates,
            totals,
            qa_pairs,
            str(tmp_path)
        )

        # Verify all estimates are in the file
        wb = load_workbook(result_path)
        ws = wb.active
        # Should have header + 3 estimates + total rows
        assert ws.max_row >= 4

    def test_export_with_english_language(self, tmp_path, sample_excel_file, mock_language_en):
        """Test Excel export with English language"""
        # Prepare test data
        estimates = [
            {"name": "Requirements", "description": "Requirements doc", "person_days": 5.0, "amount": 500.0, "reasoning": ""}
        ]

        totals = {"subtotal": 500.0, "tax": 50.0, "total": 550.0}
        qa_pairs = []

        service = ExportService()
        result_path = service.write_excel_output(
            sample_excel_file,
            estimates,
            totals,
            qa_pairs,
            str(tmp_path)
        )

        # Verify file exists
        import os
        assert os.path.exists(result_path)
